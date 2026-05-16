from pathlib import Path
import re
from typing import Any

import openpyxl


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _find_column_index(headers: list[Any], aliases: list[str]) -> int | None:
    normalized_headers = [_normalize_header(h) for h in headers]
    normalized_aliases = [_normalize_header(a) for a in aliases]
    for idx, header in enumerate(normalized_headers):
        if not header:
            continue
        for alias in normalized_aliases:
            if header == alias or alias in header:
                return idx
    return None


def _get_row_value(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def load_benchmark_cases(
    file_path: str,
    sheet_name: str | None,
    limit: int | None,
    base_dir: Path | None = None,
) -> list[dict[str, Any]]:
    path = Path(file_path)
    if not path.is_absolute():
        base = base_dir or Path.cwd()
        path = base / path

    if not path.exists():
        raise FileNotFoundError(f"Benchmark file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name and sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in workbook. Available: {wb.sheetnames}"
        )
    sheet = wb[sheet_name] if sheet_name else wb.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    col_map = {
        "question": _find_column_index(headers, ["question"]),
        "intent": _find_column_index(headers, ["intent"]),
        "tables": _find_column_index(headers, ["tables involved", "tables", "table"]),
        "columns": _find_column_index(headers, ["columns needed", "cols needed", "columns", "cols"]),
        "filters": _find_column_index(headers, ["filters/conditions", "filters", "conditions", "where"]),
        "joins": _find_column_index(headers, ["joins", "join"]),
        "sql": _find_column_index(headers, ["sql statement", "sql query", "sql", "expected sql", "actual sql"]),
    }

    if col_map["question"] is None or col_map["sql"] is None:
        raise ValueError(
            f"Required columns not found in benchmark sheet headers: {headers}"
        )

    cases: list[dict[str, Any]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        question = _get_row_value(row, col_map["question"])
        sql = _get_row_value(row, col_map["sql"])

        if not question or not str(question).strip():
            continue
        if not sql or not str(sql).strip():
            continue

        cases.append(
            {
                "row": row_idx,
                "question": str(question).strip(),
                "expected": {
                    "intent": _get_row_value(row, col_map["intent"]),
                    "tables": _get_row_value(row, col_map["tables"]),
                    "columns": _get_row_value(row, col_map["columns"]),
                    "filters": _get_row_value(row, col_map["filters"]),
                    "joins": _get_row_value(row, col_map["joins"]),
                },
                "expected_sql": str(sql).strip(),
            }
        )

        if limit is not None and limit > 0 and len(cases) >= limit:
            break

    return cases
