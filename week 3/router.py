from fastapi import APIRouter, HTTPException
import asyncio
import json
from pathlib import Path
import re
import time
from typing import Any

import openpyxl
import psycopg2

from database import test_connection
from models import PipelineResult, AIQueryRequest, SQLQueryRequest, BenchmarkRequest, SQLQueryResult
from pipeline import run_pipeline
from validator import validate_sql
from logger import log_file,logger
from executor import execute_sql

router = APIRouter(prefix="",tags=[""])


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _find_column_index(headers: list[Any], aliases: list[str]) -> int | None:
    normalized_headers = [_normalize_header(h) for h in headers]
    normalized_aliases = [_normalize_header(a) for a in aliases]
    for idx, header in enumerate(normalized_headers):
        if not header:
            continue
        for alias in normalized_aliases:
            if header == alias or alias in header:
                return idx
    return None


def _get_row_value(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def _load_benchmark_cases(
    file_path: str,
    sheet_name: str | None,
    limit: int | None,
) -> list[dict[str, Any]]:
    path = Path(file_path)
    if not path.is_absolute():
        path = Path(__file__).parent / path

    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Benchmark file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name and sheet_name not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f"Sheet '{sheet_name}' not found in workbook. Available: {wb.sheetnames}",
        )
    sheet = wb[sheet_name] if sheet_name else wb.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    col_map = {
        "question": _find_column_index(headers, ["question"]),
        "intent": _find_column_index(headers, ["Intent"]),
        "tables": _find_column_index(headers, ["Tables involved", "tables", "table"]),
        "columns": _find_column_index(headers, ["columns needed", "cols needed", "columns", "cols"]),
        "filters": _find_column_index(headers, ["filters/conditions", "filters", "conditions", "where"]),
        "joins": _find_column_index(headers, ["joins", "join"]),
        "sql": _find_column_index(headers, ["sql statement", "sql query", "sql", "expected sql", "actual sql"]),
    }

    if col_map["question"] is None or col_map["sql"] is None:
        raise HTTPException(
            status_code=400,
            detail=f"Required columns not found in benchmark sheet headers: {headers}",
        )

    cases: list[dict[str, Any]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        question = _get_row_value(row, col_map["question"])
        sql = _get_row_value(row, col_map["sql"])

        if not question or not str(question).strip():
            continue
        if not sql or not str(sql).strip():
            continue

        cases.append(
            {
                "row": row_idx,
                "question": str(question).strip(),
                "expected": {
                    "intent": _get_row_value(row, col_map["intent"]),
                    "tables": _get_row_value(row, col_map["tables"]),
                    "columns": _get_row_value(row, col_map["columns"]),
                    "filters": _get_row_value(row, col_map["filters"]),
                    "joins": _get_row_value(row, col_map["joins"]),
                },
                "expected_sql": str(sql).strip(),
            }
        )

        if limit is not None and limit > 0 and len(cases) >= limit:
            break

    return cases


def _normalize_identifier(token: str) -> str:
    cleaned = token.strip().strip("\"'`")
    lowered = cleaned.lower()
    if "*" in lowered:
        return "*"
    return lowered


def _parse_identifier_list(value: Any) -> set[str]:
    if value is None:
        return set()
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        text = str(value).strip()
        if not text or text.lower() in {"none", "n/a"}:
            return set()
        items = re.split(r"[,\n;/]+", text)

    identifiers: set[str] = set()
    for item in items:
        token = _normalize_identifier(str(item))
        if not token or token in {"none", "n/a"}:
            continue
        identifiers.add(token)
    return identifiers


def _normalize_value(value: Any) -> Any:
    if isinstance(value, dict):
        return tuple(sorted((k, _normalize_value(v)) for k, v in value.items()))
    if isinstance(value, list):
        return tuple(_normalize_value(v) for v in value)
    if isinstance(value, tuple):
        return tuple(_normalize_value(v) for v in value)
    if isinstance(value, set):
        return tuple(sorted(_normalize_value(v) for v in value))
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _normalize_rows(rows: list[dict]) -> list[tuple]:
    normalized = [
        tuple(sorted((k, _normalize_value(v)) for k, v in row.items()))
        for row in rows
    ]
    # Use repr-based ordering to avoid comparing mixed types (e.g., None vs str).
    return sorted(normalized, key=repr)


def _compare_result_sets(ai_rows: list[dict], actual_rows: list[dict]) -> bool:
    return _normalize_rows(ai_rows) == _normalize_rows(actual_rows)


def _compare_identifier_sets(expected: set[str], actual: set[str]) -> bool | None:
    if not expected:
        return None
    if not actual:
        return False
    return expected == actual


def _execute_expected_sql(sql: str) -> dict[str, Any]:
    validation = validate_sql(sql)
    if not validation.valid:
        return {
            "sql": sql,
            "row_count": 0,
            "result": [],
            "status": "failed",
            "error": validation.reason,
        }

    try:
        result = execute_sql(sql)
        return {
            "sql": sql,
            "row_count": len(result),
            "result": result,
            "status": "success",
            "error": None,
        }
    except (psycopg2.Error, Exception) as exec_err:
        return {
            "sql": sql,
            "row_count": 0,
            "result": [],
            "status": "failed",
            "error": str(exec_err),
        }


async def _run_benchmark_case(case: dict[str, Any], semaphore: asyncio.Semaphore) -> dict[str, Any]:
    async with semaphore:
        ai_task = asyncio.to_thread(run_pipeline, case["question"])
        actual_task = asyncio.to_thread(_execute_expected_sql, case["expected_sql"])
        ai_result, actual_result = await asyncio.gather(ai_task, actual_task)

    ai_tables = _parse_identifier_list(
        ai_result.decomposition.get("tables") if ai_result.decomposition else None
    )
    ai_columns = _parse_identifier_list(
        ai_result.decomposition.get("columns") if ai_result.decomposition else None
    )

    expected_tables = _parse_identifier_list(case["expected"].get("tables"))
    expected_columns = _parse_identifier_list(case["expected"].get("columns"))

    result_match = None
    if ai_result.status == "success" and actual_result["status"] == "success":
        result_match = _compare_result_sets(ai_result.result, actual_result["result"])

    return {
        "row": case["row"],
        "question": case["question"],
        "expected": case["expected"],
        "expected_sql": case["expected_sql"],
        "ai": ai_result.model_dump(),
        "actual": actual_result,
        "comparison": {
            "result_match": result_match,
            "tables_correct": _compare_identifier_sets(expected_tables, ai_tables),
            "columns_correct": _compare_identifier_sets(expected_columns, ai_columns),
        },
    }

@router.get("/health")
def health():
    db_ok = test_connection()
    return {"status": "ok", "db_connected": db_ok}

@router.post("/ai-query", response_model=PipelineResult)
def ai_query(req: AIQueryRequest):
    """Run the full Text-to-SQL pipeline for a single question."""
    if not req.question.strip():
        raise HTTPException(status_code=400, detail="Question cannot be empty")
    return run_pipeline(req.question.strip())

@router.post("/sql-query",response_model=SQLQueryResult)
def sql_query(req: SQLQueryRequest):
    """Execute the given query"""
    query = req.sql.strip()

    logger.info("=" * 60)
    logger.info(f"QUERY: {query}")

    if not query:
        raise HTTPException(status_code=400 ,detail="SQL query cannot be empty")
    
    validation = validate_sql(query)
    if not validation.valid:
        msg = f"Step 1 ✗ Validation failed: {validation.reason}"
        logger.warning(f"SQL Invalid: {msg}")
        return SQLQueryResult(
            sql=query,
            row_count=0,
            result=[],
            status="failed"
        )
    logger.info("Step 1 ✓ Validation passed")

    try:
        logger.info("Step 2 → Executing SQL against PostgreSQL")
        result = execute_sql(query)
        logger.info(f"Step 2 ✓  {len(result)} row(s) returned")
        status= 'success'
    except (psycopg2.Error, Exception) as exec_err:
        error_msg = str(exec_err)
        logger.warning(f"Step 2 ✗  Execution failed: {error_msg}")
        status = 'failed'
    return SQLQueryResult(
        sql=query,
        row_count = len(result),
        result=result,
        status=status
    )


@router.post("/benchmark")
async def benchmark(req: BenchmarkRequest):
    """
    Run the pipeline against benchmark questions from the Excel sheet,
    execute the expected SQL in parallel, and return evaluation metrics.
    """
    if req.limit is not None and req.limit <= 0:
        raise HTTPException(status_code=400, detail="limit must be a positive integer")
    if req.max_concurrency <= 0:
        raise HTTPException(status_code=400, detail="max_concurrency must be positive")

    cases = _load_benchmark_cases(req.file_path, req.sheet_name, req.limit)
    if not cases:
        raise HTTPException(status_code=400, detail="No benchmark rows found to evaluate")

    started = time.time()
    semaphore = asyncio.Semaphore(min(req.max_concurrency, len(cases)))
    tasks = [_run_benchmark_case(case, semaphore) for case in cases]
    results = await asyncio.gather(*tasks)

    total = len(results)
    ai_success = sum(1 for r in results if r["ai"]["status"] == "success")
    ai_failed = total - ai_success
    actual_success = sum(1 for r in results if r["actual"]["status"] == "success")
    actual_failed = total - actual_success

    retried = sum(1 for r in results if r["ai"]["retried"])
    retry_fixed = sum(1 for r in results if r["ai"]["retry_fixed"])

    match_count = sum(1 for r in results if r["comparison"]["result_match"] is True)
    match_evaluated = sum(
        1 for r in results if r["comparison"]["result_match"] is not None
    )

    tables_correct = sum(
        1 for r in results if r["comparison"]["tables_correct"] is True
    )
    tables_evaluated = sum(
        1 for r in results if r["comparison"]["tables_correct"] is not None
    )

    columns_correct = sum(
        1 for r in results if r["comparison"]["columns_correct"] is True
    )
    columns_evaluated = sum(
        1 for r in results if r["comparison"]["columns_correct"] is not None
    )

    latencies = [
        r["ai"]["latency_ms"]
        for r in results
        if r["ai"]["latency_ms"] is not None
    ]
    avg_latency = int(sum(latencies) / len(latencies)) if latencies else 0
    benchmark_latency = int((time.time() - started) * 1000)

    return {
        "summary": {
            "total_questions": total,
            "ai_success_count": ai_success,
            "ai_failed_count": ai_failed,
            "sql_execution_success_rate_pct": round(ai_success / total * 100, 1)
            if total
            else 0,
            "actual_success_count": actual_success,
            "actual_failed_count": actual_failed,
            "correct_sql_result_count": match_count,
            "correct_sql_result_rate_pct": round(match_count / match_evaluated * 100, 1)
            if match_evaluated
            else 0,
            "tables_correct_count": tables_correct,
            "tables_correct_rate_pct": round(tables_correct / tables_evaluated * 100, 1)
            if tables_evaluated
            else 0,
            "columns_correct_count": columns_correct,
            "columns_correct_rate_pct": round(columns_correct / columns_evaluated * 100, 1)
            if columns_evaluated
            else 0,
            "retried_count": retried,
            "retry_fixed_count": retry_fixed,
            "retry_success_rate_pct": round(retry_fixed / retried * 100, 1)
            if retried
            else 0,
            "avg_generation_latency_ms": avg_latency,
            "benchmark_latency_ms": benchmark_latency,
        },
        "results": results,
    }


@router.get("/logs")
def get_logs(limit: int = 50):
    """Return the last N log entries from today's log file."""
    if not log_file.exists():
        return {"entries": []}
    lines = log_file.read_text().splitlines()
    json_lines = [l for l in lines if l.startswith("{")]
    parsed = []
    for line in json_lines[-limit:]:
        try:
            parsed.append(json.loads(line))
        except Exception:
            pass
    return {"entries": parsed}
